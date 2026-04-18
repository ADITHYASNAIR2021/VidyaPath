#!/usr/bin/env python3
import json
import sys
from pathlib import Path

from openpyxl import load_workbook


def cell_to_text(value):
    if value is None:
        return ""
    return str(value).strip()


def main():
    if len(sys.argv) < 2:
        print(json.dumps({"error": "Missing file path"}))
        return 1

    file_path = Path(sys.argv[1])
    if not file_path.exists():
        print(json.dumps({"error": "File not found"}))
        return 1

    workbook = load_workbook(filename=str(file_path), data_only=True, read_only=True)
    sheets = {}
    preview = {"headers": [], "rows": []}

    for worksheet in workbook.worksheets:
        rows_iter = worksheet.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            sheets[worksheet.title] = []
            continue

        headers = [cell_to_text(value) for value in header_row if cell_to_text(value)]
        if not headers:
            sheets[worksheet.title] = []
            continue

        parsed_rows = []
        for row in rows_iter:
            record = {}
            has_value = False
            for index, header in enumerate(headers):
                value = cell_to_text(row[index] if index < len(row) else "")
                if value:
                    has_value = True
                record[header] = value
            if has_value:
                parsed_rows.append(record)

        sheets[worksheet.title] = parsed_rows
        if not preview["headers"]:
            preview = {"headers": headers, "rows": parsed_rows}

    print(json.dumps({"preview": preview, "sheets": sheets}))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
